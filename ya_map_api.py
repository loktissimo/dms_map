# from decimal import Decimal
from yandex_geocoder import Client
from openpyxl import load_workbook
from pprint import pprint
import json


def get_coords(address: str) -> list:

    # Yandex API connection
    client = Client("086a1aa1-6152-47a3-af40-b645a86409ae")

    # YandexAPI online request
    coordinates = client.coordinates(address)
    latitude, longitude = float(coordinates[1]), float(coordinates[0])
    return [latitude, longitude]


def coords_to_xlsx(filename):

    # Load excel workbook
    book = load_workbook(filename)

    # Get sheet
    sheet = book.active

    for i in range(2, sheet.max_row):
        if sheet['A{}'.format(i)].value:

            # Define sheet values
            # name = sheet['A{}'.format(i)].value
            # address = sheet['B{}'.format(i)].value
            # station = sheet['C{}'.format(i)].value

            # Testing without online API requests
            latitude, longitude = 55, 33

            sheet['D{}'.format(i)] = latitude
            sheet['E{}'.format(i)] = longitude

    book.save(filename=filename)


def xls_to_json(filename) -> dict:

    data = {"type": "FeatureCollection", "features": []}

    # Load excel workbook
    book = load_workbook(filename)

    # Get sheet
    sheet = book.active

    for i in range(2, sheet.max_row):
        if sheet['A{}'.format(i)].value:

            # Define sheet values
            name = sheet['A{}'.format(i)].value
            address = sheet['B{}'.format(i)].value
            station = sheet['C{}'.format(i)].value

            # Generate JSON
            geometry = {}
            geometry['type'] = 'Point'
            geometry['coordinates'] = get_coords(address)

            properties = {}
            properties['balloonContentHeader'] = name
            properties['balloonContentBody'] = address
            properties['balloonContentFooter'] = station
            properties['clusterCaption'] = name
            properties['hintContent'] = station

            poi = {}
            poi['geometry'] = geometry
            poi['properties'] = properties
            poi['type'] = 'Feature'
            poi['id'] = i-1

            data['features'].append(poi)

        # pprint(data)

    # Write json to file
    with open('data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


if __name__ == '__main__':

    # TESTING
    xls_to_json('./xls/p1_2.xlsx')

    # coords_to_xlsx('.\stom2.xlsx')
    
    # coordinates = client.coordinates("Москва, ул. Октябрьская, д. 2/4")
    # print(str(coordinates[1]), str(coordinates[0]))
    # assert coordinates == (Decimal("37.587093"), Decimal("55.733969"))

    # lat, lon = get_coords('Москва, Льва толстого 8')
    # print(lat, lon, type(lat))
    pass
